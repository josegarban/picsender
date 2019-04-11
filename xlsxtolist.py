from openpyxl import load_workbook

def get_firstrow (filename, input_sheetname = ""):
    """
    Inputs: filename and optional input sheet name. If left blank, the first row will be read.
    Objective: get values in the first row, presumably column names.
    Output: list.
    """

    wb = load_workbook(filename)
    sheetnames = wb.sheetnames
    output_list = []

    if input_sheetname == "":
        # Just get the first sheet
        ws = wb[sheetnames[0]]
        for row in ws.iter_rows(min_row=0, max_row=1, values_only=True):
            for cell in row:
                output_list.append(cell)

        print("Fieldnames found in sheet {0} in file {1}: ".format(sheetnames[0], filename), output_list)

    elif input_sheetname != "":
        if input_sheetname in sheetnames:
            ws = wb[input_sheetname]
            for row in ws.iter_rows(min_row=0, max_row=1, values_only=True):
                for cell in row:
                    output_list.append(cell)
            print("Fieldnames found in sheet {0} in file {1}: ".format(input_sheetname, filename), output_list)
        else:
            print("Sheet {0} not found in {1}.".format(input_sheetname, filename))

    return output_list

####################################################################################################

def get_otherrows (filename, input_sheetname = ""):
    """
    Inputs: filename and optional input sheet name. If left blank, the first row will be read.
    Objective: get values in the rows beside the first one, presumably values.
    Output: list.
    """

    wb = load_workbook(filename)
    sheetnames = wb.sheetnames
    output_list = []

    if input_sheetname == "":
        # Just get the first sheet
        ws = wb[sheetnames[0]]
        idx_lastrow = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=idx_lastrow, values_only=True):
            row_list = []
            for cell in row:
                row_list.append(cell)
            output_list.append(row_list)

        print("Rows found in sheet {0} in file {1}: ".format(sheetnames[0], filename), len(output_list))

    elif input_sheetname != "":
        if input_sheetname in sheetnames:
            ws = wb[input_sheetname]
            idx_lastrow = ws.max_row
            for row in ws.iter_rows(min_row=1, max_row=idx_lastrow, values_only=True):
                row_list = []
                for cell in row:
                    row_list.append(cell)
                output_list.append(row_list)

            print("Rows found in sheet {0} in file {1}: ".format(input_sheetname, filename), len(output_list))
        else:
            print("Sheet {0} not found in {1}.".format(input_sheetname, filename))

    return output_list
